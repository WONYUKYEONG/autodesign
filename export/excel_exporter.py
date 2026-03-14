"""엑셀 출력: 총괄표/내역서/일위대가표 + AI 분석 결과"""
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, numbers
from db.database import fetch_all


def export_to_excel(filepath, work_items, budget_summary, project_name=""):
    wb = Workbook()

    _create_summary_sheet(wb, budget_summary, project_name)
    _create_detail_sheet(wb, work_items)
    _create_unit_price_sheet(wb)

    wb.save(filepath)


def _style_header(ws, row, cols):
    header_font = Font(name="맑은 고딕", bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_cell(cell, is_number=False):
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.font = Font(name="맑은 고딕", size=10)
    if is_number:
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")
    else:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _create_summary_sheet(wb, s, project_name):
    ws = wb.active
    ws.title = "총괄표"
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    # 제목
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"통신선로공사 예산 총괄표"
    title_cell.font = Font(name="맑은 고딕", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    if project_name:
        ws.merge_cells('A2:D2')
        ws['A2'].value = f"프로젝트: {project_name}"
        ws['A2'].font = Font(name="맑은 고딕", size=10)

    # 헤더
    headers = ["번호", "항목", "산출근거", "금액(원)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    _style_header(ws, 4, 4)

    # 데이터
    rows_data = [
        (1, "직접 노무비", "", s.labor_cost),
        (2, "직접 재료비", "", s.material_cost),
        (3, "직접 경비", "", s.expense_cost),
        (4, "직접비 소계", "(1)+(2)+(3)", s.direct_cost),
        (5, "간접노무비", f"노무비×{12}%", s.indirect_labor),
        (6, "안전관리비", f"(노무비+재료비)×{3}%", s.safety_management),
        (7, "일반관리비", f"(노무비+경비)×{6}%", s.general_management),
        (8, "이윤", f"(노무비+일반관리비)×{5}%", s.profit),
        (9, "소계", "(4)+(5)+(6)+(7)+(8)", s.subtotal),
        (10, "부가가치세", f"소계×{10}%", s.vat),
        (11, "총 예산", "(9)+(10)", s.grand_total),
    ]

    for i, (no, name, basis, amount) in enumerate(rows_data):
        row = 5 + i
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=basis)
        ws.cell(row=row, column=4, value=amount)
        for col in range(1, 5):
            _style_cell(ws.cell(row=row, column=col), is_number=(col == 4))

    # 총 예산 행 강조
    total_row = 5 + len(rows_data) - 1
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).font = Font(name="맑은 고딕", bold=True, size=11)


def _create_detail_sheet(wb, work_items):
    ws = wb.create_sheet("내역서")
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14

    # 제목
    ws.merge_cells('A1:J1')
    ws['A1'].value = "통신선로공사 내역서"
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = ["번호", "구분", "공종/자재명", "규격", "단위", "수량",
               "노무비", "재료비", "경비", "합계"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, 10)

    total_labor = total_mat = total_exp = 0
    for i, item in enumerate(work_items):
        row = 4 + i
        type_label = "공종" if item.item_type == "work" else "자재"
        values = [
            i + 1, type_label, item.name, item.spec, item.unit,
            item.quantity, item.labor_cost, item.material_cost,
            item.expense_cost, item.total_cost,
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v)
            _style_cell(ws.cell(row=row, column=col), is_number=(col >= 6))

        total_labor += item.labor_cost
        total_mat += item.material_cost
        total_exp += item.expense_cost

    # 합계 행
    total_row = 4 + len(work_items)
    ws.cell(row=total_row, column=2, value="합계")
    ws.cell(row=total_row, column=7, value=total_labor)
    ws.cell(row=total_row, column=8, value=total_mat)
    ws.cell(row=total_row, column=9, value=total_exp)
    ws.cell(row=total_row, column=10, value=total_labor + total_mat + total_exp)
    for col in range(1, 11):
        cell = ws.cell(row=total_row, column=col)
        _style_cell(cell, is_number=(col >= 6))
        cell.font = Font(name="맑은 고딕", bold=True, size=10)


def _create_unit_price_sheet(wb):
    ws = wb.create_sheet("일위대가표")
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14

    ws.merge_cells('A1:H1')
    ws['A1'].value = "일위대가표"
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = ["번호", "코드", "공종명", "단위", "노무비", "재료비", "경비", "합계"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, 8)

    rows = fetch_all(
        """SELECT wt.code, wt.name, wt.unit,
                  up.labor_cost, up.material_cost, up.expense_cost
           FROM work_types wt
           LEFT JOIN unit_prices up ON up.work_type_code = wt.code
           ORDER BY wt.code"""
    )

    for i, r in enumerate(rows):
        row = 4 + i
        labor = r["labor_cost"] or 0
        mat = r["material_cost"] or 0
        exp = r["expense_cost"] or 0
        values = [i + 1, r["code"], r["name"], r["unit"],
                  labor, mat, exp, labor + mat + exp]
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v)
            _style_cell(ws.cell(row=row, column=col), is_number=(col >= 5))


# ──────────────────────────────────────────────
# AI 분석 결과 엑셀 출력
# ──────────────────────────────────────────────

_TYPE_KR = {
    "pole": "전주", "cable": "케이블", "terminal": "단자함",
    "house": "집/건물", "dropwire": "인입선",
}
_STATUS_KR = {
    "existing": "기존", "damaged": "파손",
    "new": "신설", "demolish": "철거",
}
_WORK_KR = {
    "install": "신설", "demolish": "철거",
    "replace": "교체", "splice": "접속",
}


def export_ai_analysis(filepath, elements, connections, ai_response=None, project_name=""):
    """AI 분석 결과를 엑셀로 출력

    Args:
        filepath: 저장 경로
        elements: list[DrawingElement] - 캔버스 요소
        connections: list[(from_id, to_id, conn_type)] - 연결 관계
        ai_response: dict - AI 원본 응답 (work_annotations 등)
        project_name: str
    """
    wb = Workbook()

    _create_elements_sheet(wb, elements, project_name)
    _create_connections_sheet(wb, elements, connections)
    if ai_response:
        _create_annotations_sheet(wb, ai_response)
    _create_summary_overview_sheet(wb, elements, connections, ai_response)

    wb.save(filepath)


def _create_elements_sheet(wb, elements, project_name):
    """요소 목록 시트"""
    ws = wb.active
    ws.title = "요소 목록"

    widths = {'A': 5, 'B': 12, 'C': 25, 'D': 10, 'E': 20, 'F': 10, 'G': 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 제목
    ws.merge_cells('A1:G1')
    title = ws['A1']
    title.value = "AI 도면 분석 - 요소 목록"
    title.font = Font(name="맑은 고딕", bold=True, size=14)
    title.alignment = Alignment(horizontal="center")

    if project_name:
        ws.merge_cells('A2:G2')
        ws['A2'].value = f"프로젝트: {project_name}"
        ws['A2'].font = Font(name="맑은 고딕", size=10)

    # 헤더
    headers = ["번호", "유형", "이름/번호", "상태", "규격", "X좌표", "Y좌표"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    _style_header(ws, 4, len(headers))

    # 데이터
    for i, elem in enumerate(elements):
        row = 5 + i
        spec = elem.properties.get("spec", "")
        values = [
            i + 1,
            _TYPE_KR.get(elem.element_type, elem.element_type),
            elem.label,
            _STATUS_KR.get(elem.status, elem.status),
            spec,
            round(elem.x),
            round(elem.y),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            _style_cell(cell, is_number=(col >= 6))

    # 요약
    summary_row = 5 + len(elements) + 1
    ws.cell(row=summary_row, column=1, value="합계")
    ws.cell(row=summary_row, column=1).font = Font(name="맑은 고딕", bold=True)

    type_counts = {}
    for elem in elements:
        name = _TYPE_KR.get(elem.element_type, elem.element_type)
        type_counts[name] = type_counts.get(name, 0) + 1
    summary_text = ", ".join(f"{k} {v}개" for k, v in type_counts.items())
    ws.cell(row=summary_row, column=2, value=summary_text)
    ws.cell(row=summary_row, column=2).font = Font(name="맑은 고딕", bold=True)


def _create_connections_sheet(wb, elements, connections):
    """연결 관계 시트"""
    ws = wb.create_sheet("연결 관계")

    widths = {'A': 5, 'B': 12, 'C': 25, 'D': 25, 'E': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:E1')
    ws['A1'].value = "요소 연결 관계"
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = ["번호", "연결 유형", "시작 요소", "끝 요소", "비고"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, len(headers))

    elem_map = {e.id: e for e in elements}

    for i, (from_id, to_id, conn_type) in enumerate(connections):
        row = 4 + i
        from_elem = elem_map.get(from_id)
        to_elem = elem_map.get(to_id)
        from_name = from_elem.display_name() if from_elem else f"ID:{from_id}"
        to_name = to_elem.display_name() if to_elem else f"ID:{to_id}"
        conn_name = _TYPE_KR.get(conn_type, conn_type)

        values = [i + 1, conn_name, from_name, to_name, ""]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            _style_cell(cell)


def _create_annotations_sheet(wb, ai_response):
    """작업 지시사항 시트"""
    annotations = ai_response.get("work_annotations", [])
    if not annotations:
        return

    ws = wb.create_sheet("작업 지시")

    widths = {'A': 5, 'B': 12, 'C': 40, 'D': 25}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:D1')
    ws['A1'].value = "AI 감지 작업 지시사항"
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = ["번호", "작업 유형", "작업 내용", "관련 요소"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, len(headers))

    for i, ann in enumerate(annotations):
        row = 4 + i
        work_type = _WORK_KR.get(ann.get("work_type", ""), ann.get("work_type", ""))
        desc = ann.get("description", "")
        related = ", ".join(ann.get("related_elements", []))

        values = [i + 1, work_type, desc, related]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            _style_cell(cell)


def _create_summary_overview_sheet(wb, elements, connections, ai_response):
    """종합 요약 시트"""
    ws = wb.create_sheet("종합 요약")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30

    ws.merge_cells('A1:C1')
    ws['A1'].value = "AI 도면 분석 종합 요약"
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    row = 3
    section_font = Font(name="맑은 고딕", bold=True, size=11)
    normal_font = Font(name="맑은 고딕", size=10)

    # 요소 현황
    ws.cell(row=row, column=1, value="▶ 요소 현황").font = section_font
    row += 1

    type_counts = {}
    status_counts = {}
    for elem in elements:
        tname = _TYPE_KR.get(elem.element_type, elem.element_type)
        type_counts[tname] = type_counts.get(tname, 0) + 1
        sname = _STATUS_KR.get(elem.status, elem.status)
        status_counts[sname] = status_counts.get(sname, 0) + 1

    for name, count in type_counts.items():
        ws.cell(row=row, column=1, value=f"  {name}").font = normal_font
        ws.cell(row=row, column=2, value=f"{count}개").font = normal_font
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="▶ 상태별 현황").font = section_font
    row += 1

    for name, count in status_counts.items():
        ws.cell(row=row, column=1, value=f"  {name}").font = normal_font
        ws.cell(row=row, column=2, value=f"{count}개").font = normal_font
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="▶ 연결 현황").font = section_font
    row += 1
    ws.cell(row=row, column=1, value=f"  총 연결 수").font = normal_font
    ws.cell(row=row, column=2, value=f"{len(connections)}개").font = normal_font

    # 작업 지시 요약
    if ai_response:
        annotations = ai_response.get("work_annotations", [])
        if annotations:
            row += 2
            ws.cell(row=row, column=1, value="▶ 작업 지시 요약").font = section_font
            row += 1
            for ann in annotations:
                desc = ann.get("description", "")
                wtype = _WORK_KR.get(ann.get("work_type", ""), "")
                ws.cell(row=row, column=1, value=f"  [{wtype}]").font = normal_font
                ws.cell(row=row, column=2, value=desc).font = normal_font
                row += 1
